import customtkinter as ctk
from tkinter import messagebox
import docx.document
import openpyxl.styles
import openpyxl.styles.cell_style
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
import docx

from Inventory import TotalInventory, Crop, Tote, LabelMaker
from gui import TimeInput, MoistureProteinInput, NewVarietyIndexWindow


class ReceivingWindow(ctk.CTkToplevel):
    def __init__(self, data:pd.DataFrame, inv_path, receiving_path, tote_label_path, loss_log_path, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.geometry('650x650')
        self.title("Receiving Form")
        self.view = ReceivingFrame(
            master=self,
            data=data,
            inv_path=inv_path,
            receiving_path=receiving_path,
            tote_label_path=tote_label_path,
            loss_log_path=loss_log_path,
            width=650,
            height=650
            )
        self.view.pack()
        self.grab_set()
        # self.transient()
        

class ReceivingFrame(ctk.CTkScrollableFrame):#FIXME scroll bar not scrolling
    def __init__(self, data:pd.DataFrame, master, inv_path, receiving_path, tote_label_path, loss_log_path, **kwargs):
        super().__init__(master, **kwargs)
        self.data = data
        self.master = master
        self.inv_path = inv_path
        self.receiving_path = receiving_path
        self.tote_label_path = tote_label_path
        self.loss_log_path = loss_log_path

        self.date_label = ctk.CTkLabel(master=self, text='Date Received (yyyy-mm-dd)', justify='right', anchor='e')
        self.date_label.grid(row=0, column=0, padx=10, pady=5)
        self.date_input = ctk.CTkEntry(master=self, placeholder_text=datetime.now().date())
        self.date_input.grid(row=0, column=1, padx=10, pady=5)

        self.time_label = ctk.CTkLabel(master=self, text='Time Received', justify='right', anchor='e')
        self.time_label.grid(row=1, column=0, padx=10, pady=5)
        self.time_input = TimeInput(master=self)
        self.time_input.grid(row=1, column=1, pady=5)

        self.crop_year_label = ctk.CTkLabel(master=self, text='Crop Year')
        self.crop_year_label.grid(row=2, column=0, pady=5)
        self.crop_year_input = ctk.CTkEntry(master=self, placeholder_text=datetime.now().date().year)
        self.crop_year_input.grid(row=2, column=1, pady=5)

        self.supplier_label = ctk.CTkLabel(master=self, text='Supplier')
        self.supplier_label.grid(row=3, column=0, pady=5)
        self.supplier_input = ctk.CTkEntry(master=self, placeholder_text='Supplier')
        self.supplier_input.grid(row=3, column=1, pady=5)

        self.grain_type_label = ctk.CTkLabel(master=self, text='Grain Type', justify='right', anchor='e')
        self.grain_type_label.grid(row=4, column=0, pady=5)
        self.g_type_input = ctk.CTkEntry(master=self, placeholder_text='Type (eg. "Wheat")')#, width=200)
        self.g_type_input.grid(row=4, column=1, padx=5, pady=5)
        self.g_variety_input = ctk.CTkEntry(master=self, placeholder_text='Variety (eg. "Butler\'s Gold")')
        self.g_variety_input.grid(row=4, column=2, padx=5, pady=5, columnspan=2)

        self.crop_id_label = ctk.CTkLabel(master=self, text='Crop ID', justify='right', anchor='e')
        self.crop_id_label.grid(row=5, column=0, pady=5)
        self.crop_id_input = ctk.CTkEntry(master=self, placeholder_text='CropID')
        self.crop_id_input.grid(row=5, column=1, pady=5)
        self.gen_crop_id_button = ctk.CTkButton(master=self, text='Generate ID', command=self.generate_id)
        self.gen_crop_id_button.grid(row=5, column=2, pady=5)

        self.parcel_id_label = ctk.CTkLabel(master=self, text='Parcel ID', justify='right', anchor='e')
        self.parcel_id_label.grid(row=6, column=0, pady=5)
        self.parcel_id_input = ctk.CTkEntry(master=self, placeholder_text='Parcel ID')
        self.parcel_id_input.grid(row=6, column=1, pady=5)

        self.organic_label = ctk.CTkLabel(master=self, text='Organic', anchor='e', justify='right')
        self.organic_label.grid(row=7, column=0, pady=5)
        self.org_var = ctk.BooleanVar()
        self.org_button = ctk.CTkRadioButton(master=self, text='Yes', variable=self.org_var, value=True)
        self.org_button.select()
        self.org_button.grid(row=7, column=1, pady=5)
        self.not_org_button = ctk.CTkRadioButton(master=self, text='No', variable=self.org_var, value=False)
        self.not_org_button.grid(row=7, column=2, pady=5)

        self.weight_label = ctk.CTkLabel(master=self, text='Weight')
        self.weight_label.grid(row=8, column=0, pady=5)
        self.weight_input = ctk.CTkEntry(master=self, placeholder_text='Weight')
        self.weight_input.grid(row=8, column=1, pady=5)
        self.num_of_totes_label = ctk.CTkLabel(master=self, text='Number of Totes\n(1) if in Grain Bin)')
        self.num_of_totes_label.grid(row=9, column=0, pady=5)
        self.num_totes_input = ctk.CTkEntry(master=self)
        self.num_totes_input.grid(row=9, column=1, pady=5)

        self.mois_prot_label = ctk.CTkLabel(master=self, text='Moisture / Protein %')
        self.mois_prot_label.grid(row=10, column=0, pady=5)
        self.m_p_input = MoistureProteinInput(master=self)
        self.m_p_input.grid(row=10, column=1, pady=5)

        self.doc_label = ctk.CTkLabel(master=self, text='Documents')
        self.doc_label.grid(row=11, column=0, pady=5)
        self.doc_input = ctk.CTkEntry(master=self, width=285, placeholder_text='Documents')
        self.doc_input.grid(row=11, column=1, pady=5, columnspan=2)

        self.mat_inspet_label = ctk.CTkLabel(master=self, text='Inspection of Raw Materials,\nIngredients, Packaging')
        self.mat_inspet_label.grid(row=12, column=0, pady=5)
        self.mat_inspect_input = ctk.CTkTextbox(master=self, width=285, height=56)
        self.mat_inspect_input.grid(row=12, column=1, pady=5, columnspan=2)

        self.trailer_insp_label = ctk.CTkLabel(master=self, text='Trailer Inspection')
        self.trailer_insp_label.grid(row=13, column=0, pady=5)
        self.trailer_insp_input = ctk.CTkEntry(master=self, width=285, placeholder_text='Condition/Cleanliness, Verify Security Seal')
        self.trailer_insp_input.grid(row=13, column=1, pady=5, columnspan=2)

        self.myco_label = ctk.CTkLabel(master=self, text='Mycotoxin Analysis\nFDA Safe Levels?')
        self.myco_label.grid(row=14, column=0, pady=5)
        self.myco_var = ctk.BooleanVar()
        self.myco_safe_button = ctk.CTkRadioButton(master=self, text='Yes', variable=self.myco_var, value=True)
        self.myco_safe_button.select()
        self.myco_safe_button.grid(row=14, column=1, pady=5)
        self.myco_unsafe_button = ctk.CTkRadioButton(master=self, text='No', variable=self.myco_var, value=False)
        self.myco_unsafe_button.grid(row=14, column=2, pady=5)

        self.rec_notes_label = ctk.CTkLabel(master=self, text='Receiving Notes\n(Parcel ID, Totes Added Automatically)')
        self.rec_notes_label.grid(row=15, column=0, pady=5)
        self.rec_notes_input = ctk.CTkTextbox(master=self, width=285, height=56)
        self.rec_notes_input.grid(row=15, column=1, pady=5, columnspan=2)

        self.inv_notes_label = ctk.CTkLabel(master=self, text='Notes for Inventory\n(Parcel ID Added Automatically)')
        self.inv_notes_label.grid(row=16, column=0, pady=5)
        self.inv_notes_input = ctk.CTkTextbox(master=self, width=285, height=56)
        self.inv_notes_input.grid(row=16, column=1, pady=5, columnspan=2)

        self.corrections_label = ctk.CTkLabel(master=self, text='Corrective Actions')
        self.corrections_label.grid(row=17, column=0, pady=5)
        self.corrections_input = ctk.CTkTextbox(master=self, width=285, height=56)
        self.corrections_input.grid(row=17, column=1, pady=5, columnspan=2)

        self.cog_label = ctk.CTkLabel(master=self, text='Cost of Goods ($/lb)')
        self.cog_label.grid(row=18, column=0, pady=5)
        self.cog_input = ctk.CTkEntry(master=self, placeholder_text='0.00')
        self.cog_input.grid(row=18, column=1, pady=5)

        self.clean_label = ctk.CTkLabel(master=self, text='Clean on Arrival?')
        self.clean_label.grid(row=19, column=0, pady=5)
        self.clean_var = ctk.BooleanVar()
        self.clean_button = ctk.CTkRadioButton(master=self, text='Yes', variable=self.clean_var, value=True)
        self.clean_button.grid(row=19, column=1, pady=5)
        self.unclean_button = ctk.CTkRadioButton(master=self, text='No', variable=self.clean_var, value=False)
        self.unclean_button.grid(row=19, column=2, pady=5)
        self.unclean_button.select()

        self.receiving_loss_label = ctk.CTkLabel(master=self, text='Receiving Loss (lbs)')
        self.receiving_loss_label.grid(row=20, column=0, pady=5)
        self.receiving_loss_input = ctk.CTkEntry(master=self, placeholder_text='0')
        self.receiving_loss_input.grid(row=20, column=1, pady=5)

        self.received_by_label = ctk.CTkLabel(master=self, text='Received By')
        self.received_by_label.grid(row=21, column=0, pady=5)
        self.received_by_input = ctk.CTkEntry(master=self, placeholder_text='Initials')
        self.received_by_input.grid(row=21, column=1, pady=5)

        self.to_inv_xl_check = ctk.CTkCheckBox(master=self, text='Write to Warehouse Inventory')
        self.to_inv_xl_check.grid(row=22, column=0, pady=5, sticky='W')
        self.to_rec_log_check = ctk.CTkCheckBox(master=self, text='Write to Receiving Log')
        self.to_rec_log_check.grid(row=23, column=0, pady=5, sticky='W')
        self.to_gohaacp_check = ctk.CTkCheckBox(master=self, text='Fill GoHAACP Form')
        self.to_gohaacp_check.grid(row=24, column=0, pady=5, sticky='W')
        self.make_labels_check = ctk.CTkCheckBox(master=self, text='Make Tote Labels')
        self.make_labels_check.grid(row=25, column=0, pady=5, sticky='W')
        self.print_labels_check = ctk.CTkCheckBox(master=self, text='Print Tote Labels')
        self.print_labels_check.grid(row=26, column=0, pady=5, sticky='W')
        self.doc_directory_check = ctk.CTkCheckBox(master=self, text='Make Directory for Receiving Documents')
        self.doc_directory_check.grid(row=27, column=0, pady=5, sticky='W')


        self.submit_button = ctk.CTkButton(master=self, text='Submit', command=self.handle_submit)
        self.submit_button.grid(row=28, column=0, pady=50)
        self.save_button = ctk.CTkButton(master=self, text='Save', command=self.handle_save)
        self.save_button.grid(row=28, column=1, pady=50)
        self.cancel_button = ctk.CTkButton(master=self, text='Cancel', command=self.handle_cancel)
        self.cancel_button.grid(row=28, column=2, pady=50)


    def generate_id(self):
        ti = TotalInventory(data=self.data)

        while self.crop_id_input.get() != '':
            self.crop_id_input.delete(first_index=0)

        if self.crop_year_input.get() == '':
            crop_year = self.crop_year_input.cget('placeholder_text')
        else: 
            crop_year = self.crop_year_input.get()
        crop = Crop(
            variety=self.g_variety_input.get(),
            supplier=self.supplier_input.get(),
            crop_year=crop_year,
            )
        self.crop_id_input.insert(index=0, string=crop.generate_crop_id(ti.get_all_crop_id()))


    def validate_input(self) -> bool: #TODO validate receiving loss
        empty_input_types = ['', None]
        not_required_empty_inputs = []
        if self.date_input.get() not in empty_input_types:
            try:
                datetime(self.date_input.get())
            except TypeError:
                messagebox.showerror(title='Invalid Date Format', message='Please Format Date Correctly')
                return False
        if self.crop_year_input.get() not in empty_input_types:
            try:
                this_year = datetime.now().year
                year = int(self.crop_year_input.get())
                if year > this_year:
                    messagebox.showerror(title='Future Grain', message='Grain cannot be received from the future... I think.')
                    return False
                if year <= this_year-10:
                    messagebox.showerror(title='Too Old', message='Too old.')
                    return False
            except ValueError:
                messagebox.showerror(title='Invalid Crop Year', message='Please enter a valid year.')
                return False
        if self.supplier_input.get() in empty_input_types:
            messagebox.showerror(title='No Supplier', message='Please enter a Supplier')
            return False
        if self.supplier_input.get() not in LabelMaker().font_info['Supplier'].keys():
            messagebox.showwarning(title='No Font Info', message=f'Font Info not found for {self.supplier_input.get()}\nDouble Check Labels Before Printing')#FIXME open font info window
        if self.g_type_input.get() in empty_input_types:
            messagebox.showerror(title='No Grain Type', message='Please enter a grain type. e.g. "Wheat", "Rye", etc...')
            return False
        if self.g_variety_input.get() in empty_input_types:
            messagebox.showerror(title='No Variety', message='Please enter a Variety Name')
            return False
        if self.g_variety_input.get() not in LabelMaker().font_info['Variety'].keys():
            messagebox.showwarning(title='No Font Info', message=f'Font info not found for {self.g_variety_input.get()}\nDouble Check Labels Before Printing') #FIXME open font info window
        if Crop(variety=self.g_variety_input.get()).get_index() == None:
            if messagebox.askokcancel(title='New Variety', message=f'{self.g_variety_input.get()} has not been assigned an Index Value.\nAssign now?'):
                nv_window = NewVarietyIndexWindow()
            return False
        if self.crop_id_input.get() in empty_input_types:
            messagebox.showerror(title='No Crop ID', message='Please enter or generate a Crop ID')
            return False
        if self.weight_input.get() in empty_input_types:
            messagebox.showerror(title='Weight', message='Please enter a weight.')
            return False
        try:
            weight = int(self.weight_input.get())
            if weight <= 0:
                messagebox.showerror(title='Invalid Weight', message='Weight must be a positive integer')
                return False
        except ValueError:
            messagebox.showerror(title='Invalid Weight', message='Weight must be a positive integer')
            return False
        if self.num_totes_input.get() in empty_input_types:
            messagebox.showerror(title='Totes', message='Please enter number of totes.')
            return False
        try:
            totes = int(self.num_totes_input.get())
            if totes <= 0:
                messagebox.showerror(title='Invalid Input', message='Number of totes must be a positive integer')
                return False
        except ValueError:
            messagebox.showerror(title='Invalid Input', message='Invalid Input for number of totes')
            return False
        if self.m_p_input.m_input.get() not in empty_input_types:
            try:
                moisture = float(self.m_p_input.m_input.get())
                if moisture <= 0:
                    messagebox.showerror(title='Moisture', message='Moisture must be positive')
                    return False
            except ValueError:
                messagebox.showerror(title='Invalid Input', message='Invalid moisture input.')
                return False
        if self.m_p_input.p_input.get() not in empty_input_types:
            try:
                protein = float(self.m_p_input.p_input.get())
                if protein < 0:
                    messagebox.showerror(title='Protein', message='Protein must be positive')
                    return False
            except ValueError:
                messagebox.showerror(title='Invalid Input', message='Invalid protein input.')
                return False
        if self.cog_input.get() not in empty_input_types:
            try:
                cog = float(self.cog_input.get())
                if cog <= 0:
                    messagebox.showerror(title='COG', message='COG must be positive')
                    return False
            except ValueError:
                messagebox.showerror(title='Invalid Input', message='Invalid COG input')
                return False
        inputs:list[ctk.CTkEntry] = [self.children[item] for item in self.children if isinstance(self.children[item], ctk.CTkEntry)]
        textboxes:list[ctk.CTkTextbox] = [self.children[item] for item in self.children if isinstance(self.children[item], ctk.CTkTextbox )]
        for item in inputs:
            if item == self.date_input or item == self.crop_year_input:
                continue
            else:
                if item.get() == '':
                    return messagebox.askyesno(title='Incomplete Fields', message='You have incomplete fields.\nContinue?')
        return True    


    def handle_checks(self, crop:Crop) -> str:
        '''Calls relevent functions based on what checkmarks are checked.
        Returns a string with the status of called functions.'''
        status = ''
        if bool(self.to_inv_xl_check.get()):
            status += f'{self.write_to_inventory(crop=crop)}\n'
        if bool(self.to_rec_log_check.get()):
            status += f'{self.write_to_receiving(crop=crop)}\n'
        if bool(self.to_gohaacp_check.get()):
            status += f'{self.write_to_gohaacp()}\n'
        if bool(self.make_labels_check.get()):
            if bool(self.print_labels_check.get()):
                status += f'{self.make_tote_labels(crop=crop, to_print=True)}\n'
            else:
                status += f'{self.make_tote_labels(crop=crop)}'
        if bool(self.doc_directory_check.get()):
            status += f'{self.make_doc_directory()}\n'
        return status


    def handle_submit(self):
        if not self.validate_input():
            return
        else:
            crop = self.make_crop()
            if crop:
                status_message = self.handle_checks(crop)
                messagebox.showinfo(title='Status', message=status_message)
                self.master.destroy()
            else: 
                messagebox.showerror(title='Aborted', message='Submit aborted due to improper formatting.')
                return


    def make_crop(self) -> Crop | None:
        if self.date_input.get() in ['', None]:
            date_rec = self.date_input.cget('placeholder_text')
        else:
            date_rec = datetime(self.date_input.get())
        if self.crop_year_input.get() in ['', None]:
            crop_year = self.crop_year_input.cget('placeholder_text')
        else:
            crop_year = self.crop_year_input.get()
        if self.m_p_input.p_input.get() in ['', None]:
            protein = 0.0
        else: 
            protein = float(self.m_p_input.p_input.get())
        if self.m_p_input.m_input.get() in ['', None]:
            moisture = 0.0
        else:
            moisture = float(self.m_p_input.m_input.get())
        if self.cog_input.get() in ['', None]:
            cog = 0.0
        else:
            cog = float(self.cog_input.get())
        try:
            crop = Crop(
                grain_type=self.g_type_input.get(), 
                variety=self.g_variety_input.get(),
                supplier=self.supplier_input.get(),
                crop_id=self.crop_id_input.get(),
                date_received=date_rec,
                crop_year=crop_year,
                moisture=moisture,
                protein=protein,
                cog=cog,
                is_org=self.org_var.get(),
                is_clean=self.clean_var.get(),
                receiving_notes=self.rec_notes_input.get('0.0', 'end').strip(),
                inventory_notes=f'{self.parcel_id_input.get()}. {self.inv_notes_input.get('0.0', 'end').strip()}',
                total_weight=int(self.weight_input.get())
            )
            tote_count = int(self.num_totes_input.get())
            ti = TotalInventory(data=self.data)
            tote_nums = crop.generate_tote_nums(all_previous_nums=ti.get_all_tote_nums(), num_to_generate=tote_count)
            crop.totes = crop.create_totes(tote_nums=tote_nums)
            return crop
        except ValueError as e:
            print(f"Error: {e}")
            return None
        

    def write_to_receiving(self, crop:Crop) -> str: #FIXME single variety ie Buckwheat will write Buckwheat, Buckwheat. Other issues? Needs further testing
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        try:
            date_to_write = f'{crop.date_received.month}.{crop.date_received.day}.{crop.date_received.year}/'
            date_to_write += f'{self.time_input.hour_input.get()}:{self.time_input.minute_input.get()}{self.time_input.am_pm.get()}'
            org_to_write = ''
            if crop.is_org:
                org_to_write = 'Y'
            else:
                org_to_write = 'N'
            myco_to_write = ''
            if self.myco_var.get():
                myco_to_write = 'FDA Safe Levels'
            else: 
                myco_to_write = 'UNSAFE LEVELS'
            wb = openpyxl.load_workbook(self.receiving_path)
            ws = wb.active
            for row in range(1, ws.max_row):
                if ws.cell(row, 1).value == None:
                    row_to_write = row
                    break
            data = {
                1: f'{crop.grain_type}, {crop.variety}',
                2: date_to_write,
                3: crop.supplier,
                4: crop.crop_id,
                5: org_to_write,
                6: crop.total_weight,
                7: self.doc_input.get(),
                8: self.mat_inspect_input.get('0.0', 'end').strip(),
                9: self.trailer_insp_input.get(),
                10: f'{self.parcel_id_input.get()}. {len(crop.totes)} x {crop.totes[0].weight}lbs totes. {crop.receiving_notes}',
                11: self.received_by_input.get(),
                12: myco_to_write,
                13: crop.moisture,
                14: crop.protein,
                15: self.corrections_input.get('0.0', 'end').strip()
            }
            for key, value in data.items():
                cell = ws.cell(row=row_to_write, column=key)
                cell.value = value
                if key == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    ws.cell(row=row_to_write, column=key).alignment = alignment

            wb.save(self.receiving_path)
            wb.close()
            return '✅ Write to Receiving Successful'
        except Exception as e:
            return f'❌ Write to Receiving Failed \n{e}\n'


    def write_to_gohaacp() -> str:
        pass #TODO

#TODO get receiving loss, update other functions to write receiving loss (write_to_receiving)
    def write_to_loss_log(self, crop:Crop) -> str:
        try:
            wb = openpyxl.load_workbook(self.loss_log_path, keep_vba=True)
            ws = wb['Receiving - Cleaning']
            grain_dv = DataValidation(type='list', formula1='"Wheat, Rye, Corn, Rice, Beans, Buckwheat"')
            ws.add_data_validation(grain_dv)
            org_dv = DataValidation(type='list', formula1='"ORGANIC, NOT ORGANIC"')
            ws.add_data_validation(org_dv)
            for row in range(1, ws.max_row):
                if ws.cell(row, 2).value == None:
                    row_to_write = row
                    break
            grain_dv.add(f'A2:A{row_to_write}')
            org_dv.add(f'F2:F{row_to_write}')
            if crop.is_org:
                org_status = 'ORGANIC'
            else:
                org_status = 'NOT ORGANIC'
            crop_data = {
                1: crop.grain_type,
                2: crop.variety,
                3: crop.crop_id,
                4: crop.date_received.strftime("%m%d%Y"),
                5: crop.supplier,
                6: org_status,
                7: crop.total_weight,
                # 8: TODO add receiving loss?
            }
            if crop.is_clean:
                crop_data.update({
                    9: 0,
                    10: crop.total_weight,
                    13: 0
                })
        except Exception as e:
            return f'❌ Write to Loss Log Failed \n{e}\n'
        #TODO Finish


    def write_to_inventory(self, crop:Crop) -> str:
        try:
            wb = openpyxl.load_workbook(self.inv_path, keep_vba=True)
            ws = wb['All']
            dv = DataValidation(type='list', formula1='"In Facility, Working, Seed Stock, Killed"')
            ws.add_data_validation(dv)
            for row in range(1, ws.max_row):
                if ws.cell(row, 2).value == None:
                    row_to_write = row
                    break
            dv.add(f'A2:A{row_to_write + len(crop.totes)}')
            org_status = 'ORGANIC'
            if not crop.is_org:
                org_status = 'Not Certified'
            cog = 0.0
            if crop.cog > 0:
                cog = crop.cog
            clean_status = ''
            if crop.is_clean:
                clean_status = 'Clean'
            for tote in crop.totes:
                tote_data = {
                    1: 'In Facility',
                    2: tote.tote_num,
                    3: tote.crop_id,
                    4: org_status,
                    5: tote.write_type_var(),
                    6: tote.supplier,
                    7: tote.date_received,
                    8: tote.protein/100,
                    9: tote.moisture/100,
                    10: cog,
                    11: tote.weight,
                    12: clean_status,
                    13: tote.weight,
                    17: tote.inv_notes
                }
                for key, value in tote_data.items():
                    cell = ws.cell(row=row_to_write, column=key)
                    if key == 1:
                        cell.alignment = Alignment(horizontal='left')
                    else:
                        cell.alignment = Alignment(horizontal='center')
                    if key in [8,9]:
                        cell.number_format = '0.00%'
                    if key == 10:
                        cell.number_format = '$ #,###0.000'
                    cell.value = value
                row_to_write += 1
            wb.save(self.inv_path)
            wb.close()
            return '✅ Write to Inventory Successful'
        except Exception as e:
            return f'❌ Write to Inventory Failed \n{e}\n'
        #TODO


    def make_tote_labels(self, crop:Crop, to_print=False) -> str:
        full_path = Path(self.tote_label_path, crop.crop_id)
        full_path.mkdir(parents=True, exist_ok=True)
        lm = LabelMaker()
        try:
            for tote in crop.totes:
                lm.make_label(tote=tote, path=full_path)
            if to_print:
                lm.print_directory(full_path)
                print_status = '✅ Print Successful'                
            return f'✅ Make Labels Successful\nSaved to "{self.tote_label_path}"\n{print_status}'
        except Exception as e:
            return f'❌ Make Labels Failed:\n{e}\n'


    def make_doc_directory(self) -> str:
        try:
            crop = self.make_crop()
        except Exception as e:
            return f'❌ Make Document Directory Failed \n{e}\n'
        base_path = Path(Path().home(), 'Team BSM Dropbox/Food Safety/Receiving/Receiving Documents')
        try:
            full_path = base_path.joinpath(base_path, str(crop.date_received.year), 'Grain & Legumes', crop.supplier, crop.variety, str(crop.date_received))
            full_path.mkdir(parents=True, exist_ok=True)
            return '✅ Create Document Directory Successful'
        except Exception as e:
            return f'❌ Make Document Directory Failed \n{e}\n'


    def handle_save(self):
        pass#TODO


    def handle_cancel(self):
        self.master.destroy()



# if __name__=='__main__':
#     def make_doc(tote:Tote):
#         doc = docx.Document()
#         print(doc.sections)
